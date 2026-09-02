import os
import sys
import subprocess
import calendar
from datetime import datetime, date, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

desktop_dir = '/Users/kushaltrivedi/Desktop/Timesheets_and_Reimbursements'
os.makedirs(desktop_dir, exist_ok=True)

template_path = os.path.join(desktop_dir, 'Hushh_Reimbursement_Tracker_Template.xlsx')
if not os.path.exists(template_path):
    template_path = '/Users/kushaltrivedi/Downloads/Hushh Reimbursement Tracker template.xlsx'

# Determine target month (previous month from today)
today = date.today()
first_of_this_month = date(today.year, today.month, 1)
last_day_prev_month = first_of_this_month - timedelta(days=1)
year = last_day_prev_month.year
month = last_day_prev_month.month

month_name = calendar.month_name[month]
num_days = calendar.monthrange(year, month)[1]

from_date = datetime(year, month, 1)
to_date = datetime(year, month, num_days)

output_filename = f"Reimbursement_Tracker_KushalTrivedi_{month_name}{year}.xlsx"
output_path = os.path.join(desktop_dir, output_filename)

print(f"Generating Timesheet for {month_name} {year} ({from_date.strftime('%Y-%m-%d')} to {to_date.strftime('%Y-%m-%d')})")
print(f"Output path: {output_path}")

# Load Workbook
wb = openpyxl.load_workbook(template_path, data_only=False)
sheet = wb.active
sheet.sheet_view.showGridLines = True

# Unmerge any old ranges in rows 10+
for rng in list(sheet.merged_cells.ranges):
    if rng.min_row >= 10:
        sheet.unmerge_cells(str(rng))

# Clear all cells from row 10 to row 160
for r in range(10, 160):
    sheet.row_dimensions[r].height = None
    for c in range(1, 16):
        cell = sheet.cell(r, c)
        cell.value = None
        cell.fill = PatternFill(fill_type=None)
        cell.border = Border()

# Styles
fill_teal_dark = PatternFill(start_color='316886', end_color='316886', fill_type='solid')
fill_teal_light = PatternFill(start_color='418AB3', end_color='418AB3', fill_type='solid')
fill_navy = PatternFill(start_color='3B4E87', end_color='3B4E87', fill_type='solid')
fill_section_hdr = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')

fill_white = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
fill_zebra = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
fill_gray_light = PatternFill(start_color='EAECEF', end_color='EAECEF', fill_type='solid')
fill_subtotal_col = PatternFill(start_color='D3D9EB', end_color='D3D9EB', fill_type='solid')
fill_grand_total = PatternFill(start_color='C5D1E8', end_color='C5D1E8', fill_type='solid')

font_header_teal = Font(name='Century Gothic', size=11, bold=True, color='FFFFFF')
font_table_hdr = Font(name='Arial', size=10, bold=True, color='FFFFFF')
font_cell_val = Font(name='Arial', size=9, bold=False, color='000000')
font_cell_link = Font(name='Arial', size=9, bold=False, color='0563C1', underline='single')
font_tot = Font(name='Arial', size=11, bold=True, color='000000')

side_thin = Side(style='thin', color='C0C0C0')
side_medium = Side(style='medium', color='3B4E87')
side_double = Side(style='double', color='3B4E87')

border_thin = Border(left=side_thin, right=side_thin, top=side_thin, bottom=side_thin)
border_total = Border(left=side_thin, right=side_thin, top=side_medium, bottom=side_double)

align_center_v = Alignment(horizontal='center', vertical='center')
align_left_v = Alignment(horizontal='left', vertical='center', wrap_text=True)
align_right_v = Alignment(horizontal='right', vertical='center')

# Header
sheet['B2'] = 'NAME'
sheet['B2'].font = font_header_teal
sheet['B2'].fill = fill_teal_dark
sheet['B2'].alignment = align_center_v
sheet['B2'].border = border_thin

sheet['C2'] = 'Kushal Trivedi'
sheet['C2'].font = Font(name='Arial', size=10, bold=True)
sheet['C2'].alignment = align_left_v
sheet['C2'].border = border_thin

sheet['F2'] = 'TIME PERIOD'
sheet['F2'].font = font_header_teal
sheet['F2'].fill = fill_teal_dark
sheet['F2'].alignment = align_center_v
sheet['F2'].border = border_thin

sheet['B3'] = 'DEPARTMENT'
sheet['B3'].font = font_header_teal
sheet['B3'].fill = fill_teal_dark
sheet['B3'].alignment = align_center_v
sheet['B3'].border = border_thin

sheet['C3'] = 'Core Agent and Engineering'
sheet['C3'].font = font_cell_val
sheet['C3'].alignment = align_left_v
sheet['C3'].border = border_thin

sheet['F3'] = 'FROM'
sheet['F3'].font = font_header_teal
sheet['F3'].fill = fill_teal_light
sheet['F3'].alignment = align_center_v
sheet['F3'].border = border_thin

sheet['G3'] = 'TO'
sheet['G3'].font = font_header_teal
sheet['G3'].fill = fill_teal_light
sheet['G3'].alignment = align_center_v
sheet['G3'].border = border_thin

sheet['B4'] = 'REPORTING'
sheet['B4'].font = font_header_teal
sheet['B4'].fill = fill_teal_dark
sheet['B4'].alignment = align_center_v
sheet['B4'].border = border_thin

sheet['C4'] = 'Manish Sainani'
sheet['C4'].font = font_cell_val
sheet['C4'].alignment = align_left_v
sheet['C4'].border = border_thin

sheet['F4'] = from_date
sheet['F4'].font = font_cell_val
sheet['F4'].number_format = 'mm/dd/yyyy'
sheet['F4'].alignment = align_center_v
sheet['F4'].border = border_thin

sheet['G4'] = to_date
sheet['G4'].font = font_cell_val
sheet['G4'].number_format = 'mm/dd/yyyy'
sheet['G4'].alignment = align_center_v
sheet['G4'].border = border_thin

sheet.column_dimensions['A'].width = 3
sheet.column_dimensions['B'].width = 16
sheet.column_dimensions['C'].width = 22
sheet.column_dimensions['D'].width = 90
sheet.column_dimensions['E'].width = 18
sheet.column_dimensions['F'].width = 15
sheet.column_dimensions['G'].width = 14
sheet.column_dimensions['H'].width = 18

# Section 1 Header
sheet.merge_cells('B8:I8')
sb8 = sheet['B8']
sb8.value = f'SECTION 1: WEEKDAY TIMESHEET & DELIVERABLES ({month_name.upper()} {year})'
sb8.font = font_header_teal
sb8.fill = fill_section_hdr
sb8.alignment = align_center_v

headers1 = [
    ('B9', 'Date'),
    ('C9', 'Payment Type'),
    ('D9', 'Deliverable Summary & Verifiable GitHub PR Link'),
    ('E9', 'Daily Allocation ($)'),
    ('F9', 'Expenses ($)'),
    ('G9', 'Misc ($)'),
    ('H9', 'Subtotal ($)')
]

for col_ref, text in headers1:
    cell = sheet[col_ref]
    cell.value = text
    cell.font = font_table_hdr
    cell.fill = fill_navy
    cell.alignment = align_center_v
    cell.border = border_thin

yearly_salary = 96969.69
monthly_salary = round(yearly_salary / 12, 2)  # $8,080.81

sheet['C11'] = 'Monthly Retainer / Base'
sheet['C11'].font = Font(name='Arial', size=9, bold=True)
sheet['C11'].fill = fill_gray_light
sheet['C11'].alignment = align_center_v
sheet['C11'].border = border_thin

sheet['D11'] = f'Yearly Base: $96,969.69 -> Monthly Base Rate: ${monthly_salary:,.2f} ($96,969.69 / 12 months)'
sheet['D11'].font = Font(name='Arial', size=9, italic=True)
sheet['D11'].fill = fill_gray_light
sheet['D11'].alignment = align_left_v
sheet['D11'].border = border_thin

sheet['H11'] = monthly_salary
sheet['H11'].font = Font(name='Arial', size=10, bold=True)
sheet['H11'].number_format = '$#,##0.00;($#,##0.00);"-";@'
sheet['H11'].alignment = align_right_v
sheet['H11'].fill = fill_subtotal_col
sheet['H11'].border = border_thin

# Fill Weekdays for month
weekdays = []
for d in range(1, num_days + 1):
    curr_dt = datetime(year, month, d)
    if curr_dt.weekday() < 5:  # Monday to Friday
        weekdays.append(curr_dt)

start_r = 13
for i, dt in enumerate(weekdays):
    r = start_r + i
    sheet.row_dimensions[r].height = 24
    row_fill = fill_zebra if i % 2 == 1 else fill_white

    sheet.cell(r, 2, value=dt).font = font_cell_val
    sheet.cell(r, 2).number_format = 'mm/dd/yyyy'
    sheet.cell(r, 2).alignment = align_center_v
    sheet.cell(r, 2).fill = row_fill
    sheet.cell(r, 2).border = border_thin

    sheet.cell(r, 3, value='Direct / Core Agent').font = font_cell_val
    sheet.cell(r, 3).alignment = align_center_v
    sheet.cell(r, 3).fill = row_fill
    sheet.cell(r, 3).border = border_thin

    display_text = f"Core Engineering & Agent Work ({dt.strftime('%b %d')})\n🔗 https://github.com/hushh-labs"
    cd = sheet.cell(r, 4, value=display_text)
    cd.hyperlink = "https://github.com/hushh-labs"
    cd.font = font_cell_link
    cd.alignment = align_left_v
    cd.fill = row_fill
    cd.border = border_thin

    sheet.cell(r, 5, value=None).fill = row_fill
    sheet.cell(r, 5).border = border_thin

    sheet.cell(r, 6, value=0).font = font_cell_val
    sheet.cell(r, 6).number_format = '$#,##0.00;($#,##0.00);"-";@'
    sheet.cell(r, 6).alignment = align_right_v
    sheet.cell(r, 6).fill = row_fill
    sheet.cell(r, 6).border = border_thin

    sheet.cell(r, 7, value=0).font = font_cell_val
    sheet.cell(r, 7).number_format = '$#,##0.00;($#,##0.00);"-";@'
    sheet.cell(r, 7).alignment = align_right_v
    sheet.cell(r, 7).fill = row_fill
    sheet.cell(r, 7).border = border_thin

    ch = sheet.cell(r, 8, value=f'=SUM(F{r}:G{r})')
    ch.font = font_cell_val
    ch.number_format = '$#,##0.00;($#,##0.00);"-";@'
    ch.alignment = align_right_v
    ch.fill = fill_subtotal_col
    ch.border = border_thin

# Section 1 Subtotal
tot_s1 = start_r + len(weekdays)
sheet.cell(tot_s1, 4, value=' Total Section 1 Monthly Retainer Pay').font = font_tot
sheet.cell(tot_s1, 4).alignment = align_right_v
sheet.cell(tot_s1, 4).fill = fill_subtotal_col
sheet.cell(tot_s1, 4).border = border_total

sheet.cell(tot_s1, 8, value='=H11').font = font_tot
sheet.cell(tot_s1, 8).number_format = '$#,##0.00;($#,##0.00);"-";@'
sheet.cell(tot_s1, 8).alignment = align_right_v
sheet.cell(tot_s1, 8).fill = fill_subtotal_col
sheet.cell(tot_s1, 8).border = border_total

wb.save(output_path)
print(f"Successfully generated monthly timesheet at {output_path}")
